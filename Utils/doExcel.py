# author:ChriPaul
# file:doExcel.py
# time:2021/02/12

from openpyxl import load_workbook


# C:\\Users\\ChriPaul\\Desktop\\\TestData.xlsx
class doExcelUtils:
    def __init__(self, fileName, sheetName, ):
        self.fileName = fileName
        self.sheetName = sheetName

    def readExcel(self):
        workBook = load_workbook(filename=self.fileName)
        sheet = workBook[self.sheetName]
        data_list = []
        for i in range(2, sheet.max_row+1):
            data_dict = {}
            for j in range(1, sheet.max_column+1):
                data_dict[sheet.cell(1, j).value] = sheet.cell(i, j).value
            data_list.append(data_dict)
        return data_list


if __name__ == '__main__':
    print(doExcelUtils(fileName="C:\\Users\\ChriPaul\\Desktop\\TestData.xlsx", sheetName="Test").readExcel())
