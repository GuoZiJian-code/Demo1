# @Author: Chris Paul
# @Time: 2021/02/18 23:26
# @File: ExcelUtil.py

import time
import pandas as pd
from openpyxl import load_workbook
from API_AUTO.utils.doConfigUtil import doConfigUtil
from API_AUTO.utils.pathUtil import pathUtils
from API_AUTO.utils.LogUtil import LogUtils
from API_AUTO.test_data.getData import getData


class do_Excel:
    def __init__(self, sheet_name,
                 workbook=load_workbook
                     (filename=pathUtils.pathApi(file_paths=["test_data", "TestData_Parameterist.xlsx"])),
                 excel_filename=pathUtils.pathApi(file_paths=["test_data", "TestData_Parameterist.xlsx"])):

        self.workbook = workbook
        self.sheet_name = sheet_name
        self.excel_filename = excel_filename

    def readExcel(self):
        mode = doConfigUtil.readConfig(filename=pathUtils.pathApi(file_paths=["ConfigFile", "TestConfig.conf"]),
                                       section="MODE", option="mode")
        sheet = self.workbook[self.sheet_name]
        result_list = []
        for i in range(2, sheet.max_row + 1):
            result_item = {}
            for j in range(1, sheet.max_column + 1):
                result_item[sheet.cell(1, j).value] = sheet.cell(i, j).value
            result_item = str(result_item)
            result_list.append(eval(result_item))

        if mode.lower() == "all":
            return self.replaceParameter(result_list=result_list,
                                         parameters=["admin_username", "admin_password", "real_name", "email", "mobile",
                                                     "password", "companyId", "account"])
        else:
            final_list = []
            for item in result_list:
                if item["caseID"] in mode:
                    final_list.append(item)
                    result_list = final_list
            return self.replaceParameter(result_list=result_list,
                                         parameters=["admin_username", "admin_password", "real_name", "email", "mobile",
                                                     "password", "companyId", "account"])

    def getCellLocationByValue(self, sheet_name, value):
        sheet = self.workbook[sheet_name]
        for i in range(1, sheet.max_row + 1):
            for j in range(1, sheet.max_column + 1):
                if sheet.cell(i, j).value == value:
                    return {"row": i, "column": j}

    def writeExcel(self, sheet_name, column_name, row_name, write_value):
        value_column_location = self.getCellLocationByValue(sheet_name=sheet_name, value=column_name)
        value_row_location = self.getCellLocationByValue(sheet_name=sheet_name, value=row_name)
        if not value_column_location:
            LogUtils().info(msg="Excel中无此列名")
            raise ValueError("Excel中无此列名")
        writeRow = value_row_location.get("row")
        writeColumn = value_column_location.get("column")
        sheet = self.workbook[sheet_name]
        sheet.cell(writeRow, writeColumn).value = write_value
        self.workbook.save(
            filename=pathUtils.pathApi(file_paths=["test_data","TestData_Parameterist.xlsx"]))

    def replaceParameter(self, result_list, parameters):
        result_list = str(result_list)
        for item in parameters:
            if result_list.find(item) != -1:
                if "test" in str(getattr(getData, item)) and "@" not in str(getattr(getData, item)):
                    result_list = result_list.replace("${" + item + "}",
                                                      "test{0}".format(time.strftime("%y%m%d%H%M%S", time.localtime())))
                    self.writeExcel(sheet_name="parameters", column_name="Value", row_name=item,
                                    write_value="test{0}".format(time.strftime("%y%m%d%H%M%S", time.localtime())))
                # 邮箱替换
                elif "test" in str(getattr(getData, item)) and "@" in str(getattr(getData, item)):
                    result_list = result_list.replace("${" + item + "}",
                                                      "{0}@test.com".format(time.strftime("%m%d%H%M%S", time.localtime())))
                    self.writeExcel(sheet_name="parameters", column_name="Value", row_name=item,
                                    write_value="{0}@test.com".format(time.strftime("%m%d%H%M%S", time.localtime())))
                # 特殊字段进行判断
                elif item == "mobile":
                    result_list = result_list.replace("${" + item + "}",
                                                      "{0}".format(
                                                          time.strftime("%y%m%d%H%M%S", time.localtime())))
                    self.writeExcel(sheet_name="parameters", column_name="Value", row_name=item,
                                    write_value="{0}".format(
                                        time.strftime("%y%m%d%H%M%S", time.localtime())))
                else:
                    result_list = result_list.replace("${" + item + "}", str(getattr(getData, item)))
                    self.writeExcel(sheet_name="parameters", column_name="Value", row_name=item,
                                    write_value=str(getattr(getData, item)))
        return eval(result_list)


if __name__ == '__main__':
    print(do_Excel(sheet_name="register").readExcel())
