# author: ChriPaul
# file: runHttpRequestTestCase_3_byExcel.py
# time: 2021 / 02 / 10
import time
import unittest
import HTMLTestReportCN
from openpyxl import load_workbook
from Day11.TestCase.httpRequestTest_2_parameterize import loginTestCase


class runHttpRequestTestCase_3_byExcel():
    if __name__ == '__main__':
        """
            使用Excel进行参数化，并请求测试TestCase
        """
        HttpRequestTestParametersWorkbookExcel = load_workbook(filename="C:\\Users\\ChriPaul\\Desktop\\TestData.xlsx")
        sheet = HttpRequestTestParametersWorkbookExcel["Test"]
        suite = unittest.TestSuite()
        for i in range(2, sheet.max_column + 2):
            #  methodName, requestUrl, requestMethod, requestData, responseExpected)
            suite.addTest(loginTestCase(methodName="testLoginApi", requestUrl=sheet.cell(i, 1).value,
                                        requestMethod=sheet.cell(i, 2).value,
                                        requestData=sheet.cell(i, 3).value, responseExpected=sheet.cell(i, 4).value))
        with open(file="D:\PythonWorkspace\Demo1\Day11\TestReport\登录以及获取用户信息接口测试{0}.html".format(
                time.strftime("%Y-%m-%d_%H-%M-%S", time.localtime())), mode="wb") as file:
            HTMLTestReportCN.HTMLTestRunner(stream=file, verbosity=2, title="登录接口测试报告",
                                            description="登录接口测试报告", tester="郭子健").run(suite)


"""
# 打开Excel
        wb = load_workbook(filename="C:\\Users\\ChriPaul\\Desktop\\TestData.xlsx")
        # 定位Sheet
        sheet = wb["Test"]  # 传表单名
        cellValue = sheet.cell(1, 1).value  # 定位并获取值
"""
