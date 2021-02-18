# author:ChriPaul
# file:doExcel.py
# time:2021/02/12
from openpyxl import load_workbook
from Utils.doConfigUtil import doConfigUtil


class doExcelUtils:
    def __init__(self, fileName, sheetName):
        self.fileName = fileName
        self.sheetName = sheetName

    def readExcel(self):
        mode = doConfigUtil().readConfig(filenames="D:\PythonWorkspace\Demo1\Day12\do_config\case.config",
                                         section="MODE", option="mode")
        workBook = load_workbook(filename=self.fileName)
        sheet = workBook[self.sheetName]
        data_list = []
        for i in range(2, sheet.max_row + 1):
            data_dict = {}
            for j in range(1, sheet.max_column + 1):
                data_dict[sheet.cell(1, j).value] = sheet.cell(i, j).value
            data_list.append(data_dict)
        if mode == "all":
            return data_list
        else:
            final_list = []
            for item in data_list:
                if item["caseID"] in mode:
                    final_list.append(item)
            return final_list


# if __name__ == '__main__':
    # print(doExcelUtils(fileName="C:\\Users\\ChriPaul\\Desktop\\TestData.xlsx", sheetName="Test").readExcel())
